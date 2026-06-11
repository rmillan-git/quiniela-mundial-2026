import io
import smtplib
import traceback
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db, settings
from models import Match, Participant, Prediction
from routes.auth import get_current_admin

router = APIRouter(prefix="/admin", tags=["admin"])

ROUND_LABELS = {
    "group_stage": "Group Stage",
    "round_of_32": "Round of 32",
    "round_of_16": "Round of 16",
    "qf": "Quarterfinals",
    "sf": "Semifinals",
    "final": "Final",
}

F_DARK_BLUE  = PatternFill("solid", fgColor="1F4E79")
F_MED_BLUE   = PatternFill("solid", fgColor="2E75B6")
F_LIGHT_BLUE = PatternFill("solid", fgColor="BDD7EE")
F_ROW        = PatternFill("solid", fgColor="D6E4F0")
F_RESULT     = PatternFill("solid", fgColor="E2EFDA")
F_EXACT      = PatternFill("solid", fgColor="C6EFCE")
F_WINNER     = PatternFill("solid", fgColor="DDEBF7")
F_WRONG      = PatternFill("solid", fgColor="FCE4D6")
F_GOLD       = PatternFill("solid", fgColor="FFD700")

WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD       = Font(bold=True)
thin       = Side(style="thin")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)


def _c(ws, row, col, value="", fill=None, font=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill = fill
    if font:  c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def build_excel(db: Session) -> bytes:
    wb = Workbook()

    participants = (
        db.query(Participant)
        .filter_by(is_approved=True, is_admin=False)
        .order_by(Participant.name)
        .all()
    )
    matches  = db.query(Match).order_by(Match.match_number).all()
    preds    = db.query(Prediction).all()
    pred_map = {(p.match_id, p.participant_id): p for p in preds}

    FIXED = 5
    p_pred_col = {p.id: FIXED + 1 + i * 2 for i, p in enumerate(participants)}
    p_pts_col  = {p.id: FIXED + 2 + i * 2 for i, p in enumerate(participants)}
    total_cols = FIXED + max(len(participants) * 2, 1)

    ws = wb.active
    ws.title = "Predictions"

    # Row 1 — title
    _merge(ws, 1, 1, 1, total_cols)
    c = ws.cell(row=1, column=1, value="⚽ Quiniela Mundial 2026 — Full Predictions")
    c.fill = F_DARK_BLUE
    c.font = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2 — column headers
    for col, lbl in [(1,"#"),(2,"Round"),(3,"Home Team"),(4,"Away Team"),(5,"Result")]:
        _c(ws, 2, col, lbl, fill=F_MED_BLUE, font=WHITE_BOLD)
    for p in participants:
        pc, ptsc = p_pred_col[p.id], p_pts_col[p.id]
        _merge(ws, 2, pc, 2, ptsc)
        _c(ws, 2, pc, p.name, fill=F_MED_BLUE, font=WHITE_BOLD)
    ws.row_dimensions[2].height = 32

    # Row 3 — sub-header Pick / Pts
    for col in range(1, FIXED + 1):
        _c(ws, 3, col, "", fill=F_LIGHT_BLUE)
    for p in participants:
        _c(ws, 3, p_pred_col[p.id], "Pick", fill=F_LIGHT_BLUE, font=BOLD)
        _c(ws, 3, p_pts_col[p.id],  "Pts",  fill=F_LIGHT_BLUE, font=BOLD)
    ws.row_dimensions[3].height = 18

    # Data rows
    row = 4
    current_round = None
    for m in matches:
        if m.round != current_round:
            current_round = m.round
            _merge(ws, row, 1, row, total_cols)
            c = ws.cell(row=row, column=1, value=ROUND_LABELS.get(m.round, m.round))
            c.fill = F_DARK_BLUE; c.font = WHITE_BOLD
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = BORDER
            ws.row_dimensions[row].height = 20
            row += 1

        home   = m.home_team.name if m.home_team else (m.home_team_placeholder or "TBD")
        away   = m.away_team.name if m.away_team else (m.away_team_placeholder or "TBD")
        result = f"{m.home_score}-{m.away_score}" if m.is_finished else "-"

        _c(ws, row, 1, m.match_number, fill=F_ROW, font=BOLD)
        _c(ws, row, 2, ROUND_LABELS.get(m.round, m.round), fill=F_ROW)
        _c(ws, row, 3, home, fill=F_ROW, align="left")
        _c(ws, row, 4, away, fill=F_ROW, align="left")
        _c(ws, row, 5, result, fill=F_RESULT, font=BOLD)

        for p in participants:
            pred = pred_map.get((m.id, p.id))
            pc   = p_pred_col[p.id]
            ptsc = p_pts_col[p.id]
            if pred:
                pick = f"{pred.home_score}-{pred.away_score}"
                pts  = pred.points
                if m.is_finished and pts is not None:
                    pf = F_EXACT if pts >= 9 else (F_WINNER if pts >= 5 else F_WRONG)
                else:
                    pf = None
                _c(ws, row, pc,   pick, fill=pf)
                _c(ws, row, ptsc, pts if pts is not None else "", fill=pf, font=BOLD if pts else None)
            else:
                _c(ws, row, pc,   "-")
                _c(ws, row, ptsc, "")
        row += 1

    # Totals row
    row += 1
    _merge(ws, row, 1, row, FIXED)
    c = ws.cell(row=row, column=1, value="TOTAL POINTS")
    c.fill = F_DARK_BLUE; c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = BORDER

    for p in participants:
        total = sum(
            pred_map[(m.id, p.id)].points or 0
            for m in matches
            if (m.id, p.id) in pred_map and pred_map[(m.id, p.id)].points is not None
        )
        pc, ptsc = p_pred_col[p.id], p_pts_col[p.id]
        _merge(ws, row, pc, row, ptsc)
        c = ws.cell(row=row, column=pc, value=total)
        c.fill = F_GOLD; c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 24

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 8
    for p in participants:
        ws.column_dimensions[get_column_letter(p_pred_col[p.id])].width = 8
        ws.column_dimensions[get_column_letter(p_pts_col[p.id])].width  = 5
    ws.freeze_panes = "F4"

    # ── Sheet 2: Leaderboard ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Leaderboard")
    hdrs = ["Rank","Name","Email","Group Stage","Round of 32","Round of 16",
            "Quarterfinals","Semifinals","Final","TOTAL"]
    for col, h in enumerate(hdrs, 1):
        _c(ws2, 1, col, h, fill=F_DARK_BLUE, font=WHITE_BOLD)
    ws2.row_dimensions[1].height = 22

    scores = []
    for p in participants:
        by_round = {}
        for m in matches:
            pred = pred_map.get((m.id, p.id))
            if pred and pred.points is not None:
                by_round[m.round] = by_round.get(m.round, 0) + pred.points
        scores.append({"name": p.name, "email": p.email, "by_round": by_round,
                       "total": sum(by_round.values())})
    scores.sort(key=lambda x: x["total"], reverse=True)

    for i, s in enumerate(scores, 1):
        br = s["by_round"]
        vals = [i, s["name"], s["email"],
                br.get("group_stage",0), br.get("round_of_32",0), br.get("round_of_16",0),
                br.get("qf",0), br.get("sf",0), br.get("final",0), s["total"]]
        for col, v in enumerate(vals, 1):
            fill = F_GOLD if col == 10 else None
            font = Font(bold=True) if col in (1, 10) else None
            _c(ws2, i + 1, col, v, fill=fill, font=font)

    for col, w in zip(range(1, 11), [6, 24, 28, 13, 13, 13, 14, 12, 8, 10]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db), _: Participant = Depends(get_current_admin)):
    try:
        data = build_excel(db)
    except Exception as e:
        raise HTTPException(500, detail=f"{type(e).__name__}: {e}\n{traceback.format_exc()}")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quiniela-mundial-2026.xlsx"},
    )


@router.post("/send-report")
def send_report(db: Session = Depends(get_db), _: Participant = Depends(get_current_admin)):
    if not settings.gmail_user or not settings.gmail_app_password:
        raise HTTPException(500, "Gmail credentials not configured")

    try:
        excel_data = build_excel(db)
    except Exception as e:
        raise HTTPException(500, detail=f"Excel build failed: {e}")

    participants = (
        db.query(Participant)
        .filter_by(is_approved=True)
        .all()
    )
    recipients = [p.email for p in participants if p.email]
    today = datetime.now().strftime("%B %d, %Y")
    filename = f"quiniela-mundial-2026-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # Build leaderboard summary for email body
    preds = db.query(Prediction).all()
    pred_map = {(p.match_id, p.participant_id): p for p in preds}
    matches = db.query(Match).all()
    scores = []
    for p in participants:
        if p.is_admin: continue
        total = sum(
            pred_map[(m.id, p.id)].points or 0
            for m in matches
            if (m.id, p.id) in pred_map and pred_map[(m.id, p.id)].points is not None
        )
        scores.append((p.name, total))
    scores.sort(key=lambda x: x[1], reverse=True)

    standings_html = "".join(
        f"<tr><td>{i}</td><td>{name}</td><td><b>{pts}</b></td></tr>"
        for i, (name, pts) in enumerate(scores, 1)
    )

    html_body = f"""
    <html><body>
    <h2>⚽ Quiniela Mundial 2026 — Daily Report</h2>
    <p>{today}</p>
    <h3>📊 Current Standings</h3>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse">
      <tr style="background:#1F4E79;color:white"><th>#</th><th>Name</th><th>Points</th></tr>
      {standings_html}
    </table>
    <p>See the full predictions and results in the attached Excel file.</p>
    <p>🌍 <a href="https://quiniela-frontend-l8j1.onrender.com">View Leaderboard</a></p>
    </body></html>
    """

    sent = 0
    failed = []
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(settings.gmail_user, settings.gmail_app_password)

        for recipient in recipients:
            try:
                msg = MIMEMultipart()
                msg["From"] = settings.gmail_user
                msg["To"] = recipient
                msg["Subject"] = f"⚽ Quiniela Mundial 2026 — Daily Report {today}"
                msg.attach(MIMEText(html_body, "html"))

                part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                part.set_payload(excel_data)
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(part)

                server.sendmail(settings.gmail_user, recipient, msg.as_string())
                sent += 1
            except Exception as e:
                failed.append(f"{recipient}: {e}")

        server.quit()
    except Exception as e:
        raise HTTPException(500, detail=f"SMTP error: {e}")

    return {"sent": sent, "failed": failed, "recipients": recipients}
